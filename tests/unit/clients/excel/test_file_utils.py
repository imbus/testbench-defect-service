from datetime import datetime
from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl
import pandas as pd
import pytest

from testbench_defect_service.clients.excel.config import (
    ExcelDefectClientConfig,
    UserDefiendAttributes,
)
from testbench_defect_service.clients.excel.file_utils import (
    _apply_boolean_udf_write_mapping,
    map_boolean_values,
    read_data_frame_from_file_path,
    write_defect_data,
    write_defect_data_to_csv,
    write_defect_data_to_excel,
)
from testbench_defect_service.clients.excel.utils import is_blank_row
from testbench_defect_service.models.defects import Protocol, SyncContext, ValueType

_MODULE = "testbench_defect_service.clients.excel.file_utils"


@pytest.fixture
def config() -> ExcelDefectClientConfig:
    return ExcelDefectClientConfig(
        system_name="Excel",
        excel_file_path=Path(),
        file_type=".xlsx",
        simple_date_format="yyyy-MM-dd",
        defects_data_header_line=1,
        defects_data_starting_line=2,
        separator=",",
        id_column_no=1,
        title_column_no=2,
        references_column_no=0,
        discoverer_column_no=0,
        lastedit_column_no=0,
        description_column_no=0,
        references_separator=";",
        id_prefix="D-",
        defect_id_starting_value="1",
        defect_id_digit_numbers=4,
        control_fields=[],
    )


@pytest.fixture
def sync_context() -> SyncContext:
    return SyncContext()


@pytest.fixture
def df() -> pd.DataFrame:
    return pd.DataFrame({"id": ["D-1"], "title": ["Bug"]})


@pytest.fixture
def header() -> dict[int, str]:
    return {1: "Defect ID", 2: "Title"}


@pytest.mark.unit
class TestWriteDefectDataToExcel:
    def test_new_file_uses_write_mode(
        self,
        tmp_path: Path,
        config: ExcelDefectClientConfig,
        sync_context: SyncContext,
        df: pd.DataFrame,
        header: dict[int, str],
    ):
        defect_path = tmp_path / "output.xlsx"
        assert not defect_path.exists()

        with (
            patch(f"{_MODULE}.pd.ExcelWriter") as mock_excel_writer,
            patch(f"{_MODULE}.resolve_visible_sheet_name", return_value="Sheet1"),
            patch(f"{_MODULE}._clear_stale_excel_rows"),
            patch.object(pd.DataFrame, "to_excel"),
        ):
            write_defect_data_to_excel(sync_context, defect_path, config, header, df)

        mock_excel_writer.assert_called_once()
        call_kwargs = mock_excel_writer.call_args.kwargs
        assert call_kwargs["engine"] == "openpyxl"
        assert "mode" not in call_kwargs
        assert "if_sheet_exists" not in call_kwargs

    def test_existing_xlsx_uses_append_mode_with_overlay(
        self,
        tmp_path: Path,
        config: ExcelDefectClientConfig,
        sync_context: SyncContext,
        df: pd.DataFrame,
        header: dict[int, str],
    ):
        defect_path = tmp_path / "existing.xlsx"
        defect_path.write_bytes(b"")

        with (
            patch(f"{_MODULE}.pd.ExcelWriter") as mock_excel_writer,
            patch(f"{_MODULE}.resolve_visible_sheet_name", return_value="Sheet1"),
            patch(f"{_MODULE}._clear_stale_excel_rows"),
            patch.object(pd.DataFrame, "to_excel"),
        ):
            write_defect_data_to_excel(sync_context, defect_path, config, header, df)

        mock_excel_writer.assert_called_once()
        call_kwargs = mock_excel_writer.call_args.kwargs
        assert call_kwargs["engine"] == "openpyxl"
        assert call_kwargs["mode"] == "a"
        assert call_kwargs["if_sheet_exists"] == "overlay"

    def test_xls_target_raises_without_opening_writer(
        self,
        tmp_path: Path,
        config: ExcelDefectClientConfig,
        sync_context: SyncContext,
        df: pd.DataFrame,
        header: dict[int, str],
    ):
        defect_path = tmp_path / "existing.xls"
        original_bytes = b"\xd0\xcf\x11\xe0legacy-xls-content"
        defect_path.write_bytes(original_bytes)

        with (
            patch(f"{_MODULE}.pd.ExcelWriter") as mock_excel_writer,
            pytest.raises(ValueError, match=r"\.xls"),
        ):
            write_defect_data_to_excel(sync_context, defect_path, config, header, df)

        mock_excel_writer.assert_not_called()
        assert defect_path.read_bytes() == original_bytes

    def test_none_column_positions_returns_early_without_writing(
        self,
        tmp_path: Path,
        config: ExcelDefectClientConfig,
        sync_context: SyncContext,
        df: pd.DataFrame,
        header: dict[int, str],
    ):
        defect_path = tmp_path / "output.xlsx"

        with (
            patch(f"{_MODULE}.get_column_mapping_for_config", return_value=None),
            patch(f"{_MODULE}.pd.ExcelWriter") as mock_excel_writer,
        ):
            write_defect_data_to_excel(sync_context, defect_path, config, header, df)

        mock_excel_writer.assert_not_called()

    def test_empty_column_positions_returns_early_without_writing(
        self,
        tmp_path: Path,
        config: ExcelDefectClientConfig,
        sync_context: SyncContext,
        df: pd.DataFrame,
        header: dict[int, str],
    ):
        defect_path = tmp_path / "output.xlsx"

        with (
            patch(f"{_MODULE}.get_column_mapping_for_config", return_value={}),
            patch(f"{_MODULE}.pd.ExcelWriter") as mock_excel_writer,
        ):
            write_defect_data_to_excel(sync_context, defect_path, config, header, df)

        mock_excel_writer.assert_not_called()


@pytest.mark.unit
class TestReadDataFrameFromFilePath:
    @pytest.fixture
    def file_path(self, tmp_path: Path) -> Path:
        path = tmp_path / "data.xlsx"
        path.touch()
        return path

    def test_returns_empty_dataframe_when_column_mapping_is_none(
        self,
        file_path: Path,
        config: ExcelDefectClientConfig,
        sync_context: SyncContext,
    ):
        loaded_df = pd.DataFrame({"col_a": ["x", "y"], "col_b": ["1", "2"]})

        with (
            patch(f"{_MODULE}._load_dataframe", return_value=loaded_df),
            patch(f"{_MODULE}.get_column_mapping_for_config", return_value=None),
        ):
            result = read_data_frame_from_file_path(file_path, config, sync_context)

        assert result.empty
        assert list(result.columns) == list(loaded_df.columns)

    def test_returns_mapped_dataframe_on_happy_path(
        self,
        file_path: Path,
        config: ExcelDefectClientConfig,
        sync_context: SyncContext,
    ):
        loaded_df = pd.DataFrame({"col_a": ["x"], "col_b": ["1"]})
        expected_df = pd.DataFrame({"id": ["x"], "title": ["1"]})
        valid_mapping = {0: ["id"], 1: ["title"]}

        with (
            patch(f"{_MODULE}._load_dataframe", return_value=loaded_df),
            patch(f"{_MODULE}.get_column_mapping_for_config", return_value=valid_mapping),
            patch(f"{_MODULE}._validate_column_mapping", return_value=valid_mapping),
            patch(f"{_MODULE}.map_boolean_values"),
            patch(f"{_MODULE}._apply_column_mapping", return_value=expected_df),
            patch(f"{_MODULE}._require_mapped_columns"),
        ):
            result = read_data_frame_from_file_path(file_path, config, sync_context)

        assert result is expected_df

    def test_map_boolean_values_is_called_with_mapped_df(
        self,
        file_path: Path,
        config: ExcelDefectClientConfig,
        sync_context: SyncContext,
    ):
        loaded_df = pd.DataFrame({"Header text": ["x"]})
        mapped_df = pd.DataFrame({"id": ["x"]})
        valid_mapping = {0: ["id"]}

        with (
            patch(f"{_MODULE}._load_dataframe", return_value=loaded_df),
            patch(f"{_MODULE}.get_column_mapping_for_config", return_value=valid_mapping),
            patch(f"{_MODULE}._validate_column_mapping", return_value=valid_mapping),
            patch(f"{_MODULE}.map_boolean_values") as mock_map_bool,
            patch(f"{_MODULE}._apply_column_mapping", return_value=mapped_df),
            patch(f"{_MODULE}._require_mapped_columns"),
        ):
            read_data_frame_from_file_path(file_path, config, sync_context)

        mock_map_bool.assert_called_once_with(config, mapped_df)

    def test_required_columns_are_checked_against_the_mapped_frame(
        self,
        file_path: Path,
        config: ExcelDefectClientConfig,
        sync_context: SyncContext,
    ):
        loaded_df = pd.DataFrame({"id": ["x"]})
        mapped_df = pd.DataFrame({"id": ["x"]})
        valid_mapping = {0: ["id"]}

        with (
            patch(f"{_MODULE}._load_dataframe", return_value=loaded_df),
            patch(f"{_MODULE}.get_column_mapping_for_config", return_value=valid_mapping),
            patch(f"{_MODULE}._validate_column_mapping", return_value=valid_mapping),
            patch(f"{_MODULE}.map_boolean_values"),
            patch(f"{_MODULE}._apply_column_mapping", return_value=mapped_df),
            patch(f"{_MODULE}._require_mapped_columns") as mock_require_columns,
        ):
            read_data_frame_from_file_path(file_path, config, sync_context)

        mock_require_columns.assert_called_once_with(mapped_df, file_path)

    def test_propagates_missing_required_column_error(
        self,
        file_path: Path,
        config: ExcelDefectClientConfig,
        sync_context: SyncContext,
    ):
        loaded_df = pd.DataFrame({"id": [""]})
        mapped_df = pd.DataFrame({"id": [""]})
        valid_mapping = {0: ["id"]}

        with (
            patch(f"{_MODULE}._load_dataframe", return_value=loaded_df),
            patch(f"{_MODULE}.get_column_mapping_for_config", return_value=valid_mapping),
            patch(f"{_MODULE}._validate_column_mapping", return_value=valid_mapping),
            patch(f"{_MODULE}.map_boolean_values"),
            patch(f"{_MODULE}._apply_column_mapping", return_value=mapped_df),
            patch(
                f"{_MODULE}._require_mapped_columns",
                side_effect=ValueError("column not found"),
            ),
            pytest.raises(ValueError, match="column not found"),
        ):
            read_data_frame_from_file_path(file_path, config, sync_context)

    def test_protocol_is_passed_to_get_column_mapping(
        self,
        file_path: Path,
        config: ExcelDefectClientConfig,
        sync_context: SyncContext,
    ):
        protocol = MagicMock(spec=Protocol)
        loaded_df = pd.DataFrame({"id": ["x"]})
        valid_mapping = {0: ["id"]}
        mapped_df = pd.DataFrame({"id": ["x"]})

        with (
            patch(f"{_MODULE}._load_dataframe", return_value=loaded_df),
            patch(
                f"{_MODULE}.get_column_mapping_for_config", return_value=valid_mapping
            ) as mock_get_mapping,
            patch(f"{_MODULE}._validate_column_mapping", return_value=valid_mapping),
            patch(f"{_MODULE}.map_boolean_values"),
            patch(f"{_MODULE}._apply_column_mapping", return_value=mapped_df),
            patch(f"{_MODULE}._require_mapped_columns"),
        ):
            read_data_frame_from_file_path(file_path, config, sync_context, protocol=protocol)

        mock_get_mapping.assert_called_once_with(config, sync_context, protocol)


# ---------------------------------------------------------------------------
# Format-specific behavior of the delimited writer
# ---------------------------------------------------------------------------


@pytest.fixture
def two_row_df() -> pd.DataFrame:
    return pd.DataFrame({"id": ["D-1", "D-2"], "title": ["Bug", "New"]})


@pytest.mark.unit
class TestWriteDefectDataToCsvFormats:
    def test_semicolon_separator_preserves_file_structure(
        self,
        tmp_path: Path,
        config: ExcelDefectClientConfig,
        sync_context: SyncContext,
        two_row_df: pd.DataFrame,
        header: dict[int, str],
    ):
        defect_path = tmp_path / "defects.csv"
        defect_path.write_text("id;title\nD-1;Bug\n", encoding="utf-8")
        semicolon_config = config.model_copy(update={"separator": ";"})

        write_defect_data_to_csv(
            sync_context, defect_path, semicolon_config, {1: "id", 2: "title"}, two_row_df
        )

        lines = defect_path.read_text(encoding="utf-8").splitlines()
        assert lines[0] == "id;title"
        assert lines[1] == "D-1;Bug"
        assert lines[2] == "D-2;New"

    def test_unmapped_na_like_values_are_preserved(
        self,
        tmp_path: Path,
        config: ExcelDefectClientConfig,
        sync_context: SyncContext,
        two_row_df: pd.DataFrame,
    ):
        defect_path = tmp_path / "defects.csv"
        defect_path.write_text("id,title,notes\nD-1,Bug,N/A\n", encoding="utf-8")

        write_defect_data_to_csv(
            sync_context, defect_path, config, {1: "id", 2: "title"}, two_row_df
        )

        content = defect_path.read_text(encoding="utf-8")
        assert "N/A" in content

    def test_windows_1252_file_is_written_back_in_same_encoding(
        self,
        tmp_path: Path,
        config: ExcelDefectClientConfig,
        sync_context: SyncContext,
    ):
        defect_path = tmp_path / "defects.csv"
        defect_path.write_bytes("id,title\nD-1,Käfer\n".encode("windows-1252"))
        df = pd.DataFrame({"id": ["D-1", "D-2"], "title": ["Käfer", "New"]})

        write_defect_data_to_csv(sync_context, defect_path, config, {1: "id", 2: "title"}, df)

        content = defect_path.read_bytes().decode("windows-1252")
        assert "Käfer" in content
        assert "D-2" in content

    def test_missing_separator_defaults_to_comma(
        self,
        tmp_path: Path,
        config: ExcelDefectClientConfig,
        sync_context: SyncContext,
        two_row_df: pd.DataFrame,
    ):
        defect_path = tmp_path / "defects.csv"
        defect_path.write_text("id,title\nD-1,Bug\n", encoding="utf-8")
        no_separator_config = config.model_copy(update={"separator": None})

        write_defect_data_to_csv(
            sync_context, defect_path, no_separator_config, {1: "id", 2: "title"}, two_row_df
        )

        lines = defect_path.read_text(encoding="utf-8").splitlines()
        assert lines[0] == "id,title"
        assert lines[2] == "D-2,New"

    def test_txt_file_keeps_unmapped_columns(
        self,
        tmp_path: Path,
        config: ExcelDefectClientConfig,
        sync_context: SyncContext,
        two_row_df: pd.DataFrame,
    ):
        defect_path = tmp_path / "defects.txt"
        defect_path.write_text("id,title,notes\nD-1,Bug,keep-me\n", encoding="utf-8")

        write_defect_data_to_csv(
            sync_context, defect_path, config, {1: "id", 2: "title"}, two_row_df
        )

        content = defect_path.read_text(encoding="utf-8")
        assert "keep-me" in content
        assert "D-2" in content

    def test_stale_trailing_row_is_cleared_after_delete(
        self,
        tmp_path: Path,
        config: ExcelDefectClientConfig,
        sync_context: SyncContext,
        two_row_df: pd.DataFrame,
    ):
        defect_path = tmp_path / "defects.csv"
        defect_path.write_text(
            "id,title\nD-1,Bug\nD-2,New\nD-3,Deleted\n",
            encoding="utf-8",
        )

        write_defect_data_to_csv(
            sync_context, defect_path, config, {1: "id", 2: "title"}, two_row_df
        )

        content = defect_path.read_text(encoding="utf-8")
        assert "D-3" not in content
        assert "Deleted" not in content


@pytest.mark.unit
class TestWriteDefectDataDispatch:
    def test_tsv_is_written_tab_separated(
        self,
        tmp_path: Path,
        config: ExcelDefectClientConfig,
        sync_context: SyncContext,
        two_row_df: pd.DataFrame,
    ):
        defect_path = tmp_path / "defects.tsv"
        defect_path.write_text("id\ttitle\nD-1\tBug\n", encoding="utf-8")

        write_defect_data(sync_context, defect_path, config, {1: "id", 2: "title"}, two_row_df)

        lines = defect_path.read_text(encoding="utf-8").splitlines()
        assert lines[0] == "id\ttitle"
        assert lines[2] == "D-2\tNew"

    def test_unsupported_suffix_raises_value_error(
        self,
        tmp_path: Path,
        config: ExcelDefectClientConfig,
        sync_context: SyncContext,
        two_row_df: pd.DataFrame,
    ):
        defect_path = tmp_path / "defects.ods"
        defect_path.write_text("irrelevant", encoding="utf-8")

        with pytest.raises(ValueError, match="Unsupported file format"):
            write_defect_data(sync_context, defect_path, config, {1: "id", 2: "title"}, two_row_df)


@pytest.mark.unit
class TestWriteDefectDataToExcelStaleRows:
    def test_stale_trailing_row_is_cleared_after_delete(
        self,
        tmp_path: Path,
        config: ExcelDefectClientConfig,
        sync_context: SyncContext,
        two_row_df: pd.DataFrame,
    ):
        defect_path = tmp_path / "defects.xlsx"
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        assert worksheet is not None
        worksheet.title = "Sheet1"
        worksheet.append(["id", "title"])
        worksheet.append(["D-1", "Bug"])
        worksheet.append(["D-2", "New"])
        worksheet.append(["D-3", "Deleted"])
        workbook.save(defect_path)

        write_defect_data_to_excel(
            sync_context, defect_path, config, {1: "id", 2: "title"}, two_row_df
        )

        saved = openpyxl.load_workbook(defect_path)
        sheet = saved["Sheet1"]
        assert sheet["A4"].value in (None, "")
        assert sheet["B4"].value in (None, "")
        assert sheet["A3"].value == "D-2"


# ---------------------------------------------------------------------------
# Format-specific behavior of the readers
# ---------------------------------------------------------------------------


@pytest.mark.unit
class TestLoadDataFrameFormats:
    def test_excel_datetime_cells_use_configured_date_format(
        self,
        tmp_path: Path,
        config: ExcelDefectClientConfig,
        sync_context: SyncContext,
    ):
        defect_path = tmp_path / "defects.xlsx"
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        assert worksheet is not None
        worksheet.title = "Sheet1"
        worksheet.append(["id", "title", "edited"])
        worksheet.append(["D-1", "Bug", datetime(2024, 1, 5)])  # noqa: DTZ001 - Excel cells are naive
        workbook.save(defect_path)
        date_config = config.model_copy(update={"lastedit_column_no": 3})

        df = read_data_frame_from_file_path(defect_path, date_config, sync_context)

        assert df["lastEdited"].iloc[0] == "2024-01-05"

    def test_boolean_udf_maps_when_header_differs_from_udf_name(
        self,
        tmp_path: Path,
        config: ExcelDefectClientConfig,
        sync_context: SyncContext,
    ):
        defect_path = tmp_path / "defects.csv"
        defect_path.write_text("id,title,Fixed?\nD-1,Bug,yes\n", encoding="utf-8")
        udf_config = config.model_copy(
            update={
                "udfs": [
                    UserDefiendAttributes(
                        name="isFixed",
                        column=3,
                        type=ValueType.BOOLEAN,
                        trueValue="yes",
                        falseValue="no",
                    )
                ]
            }
        )

        df = read_data_frame_from_file_path(defect_path, udf_config, sync_context)

        assert df["isFixed"].iloc[0] == "true"


@pytest.mark.unit
class TestEmptyRowTolerance:
    """A separator-only row is layout, not data. It must not fail the whole file.

    Truly blank lines are already dropped by pandas (`skip_blank_lines=True`);
    the breaking case is `,,` which Excel, LibreOffice and this project's own
    CSV writer all emit.
    """

    def _write(self, tmp_path: Path, body: str) -> Path:
        file_path = tmp_path / "defects.csv"
        file_path.write_text(f"id,title\n{body}", encoding="utf-8")
        return file_path

    def test_empty_row_is_kept_in_the_frame_and_does_not_raise(
        self,
        tmp_path: Path,
        config: ExcelDefectClientConfig,
        sync_context: SyncContext,
    ):
        file_path = self._write(tmp_path, "D-1,First\n,\nD-2,Second\n")

        df = read_data_frame_from_file_path(file_path, config, sync_context, Protocol())

        # The row is preserved: the frame stays positionally aligned with the file.
        assert df["id"].tolist() == ["D-1", "", "D-2"]

    def test_several_empty_rows_do_not_trip_the_uniqueness_check(
        self,
        tmp_path: Path,
        config: ExcelDefectClientConfig,
        sync_context: SyncContext,
    ):
        file_path = self._write(tmp_path, "D-1,First\n,\n,\nD-2,Second\n")

        df = read_data_frame_from_file_path(file_path, config, sync_context, Protocol())

        assert df["id"].tolist() == ["D-1", "", "", "D-2"]


@pytest.mark.unit
class TestUnusableRowTolerance:
    """A row the service cannot identify is one bad row, not a bad file.

    Failing the read took the whole file down on every path, update and delete included, so
    the only way out was editing the file by hand. The rows stay in the frame: dropping them
    would shift every row below up on the next write and erase what the user has to repair.
    """

    def _write(self, tmp_path: Path, body: str) -> Path:
        file_path = tmp_path / "defects.csv"
        file_path.write_text(f"id,title\n{body}", encoding="utf-8")
        return file_path

    def test_row_with_blank_id_but_other_content_does_not_fail_the_file(
        self,
        tmp_path: Path,
        config: ExcelDefectClientConfig,
        sync_context: SyncContext,
    ):
        file_path = self._write(tmp_path, "D-1,First\n,Only a title\nD-2,Second\n")

        df = read_data_frame_from_file_path(file_path, config, sync_context, Protocol())

        assert df["id"].tolist() == ["D-1", "", "D-2"]

    def test_duplicate_ids_do_not_fail_the_file(
        self,
        tmp_path: Path,
        config: ExcelDefectClientConfig,
        sync_context: SyncContext,
    ):
        """The rows stay positionally aligned with the file, which is what keeps the next write
        from shifting the rows below them up and erasing what the user has to repair."""
        file_path = self._write(tmp_path, "D-1,First\n,\nD-1,Duplicate\n")

        df = read_data_frame_from_file_path(file_path, config, sync_context, Protocol())

        assert df["id"].tolist() == ["D-1", "", "D-1"]

    def test_missing_id_column_still_fails_the_file(
        self,
        tmp_path: Path,
        config: ExcelDefectClientConfig,
        sync_context: SyncContext,
    ):
        """An unmapped id column is a configuration fault, not a row fault - nothing in the
        file can be identified, so there is no partial result worth returning."""
        file_path = self._write(tmp_path, "D-1,First\n")
        unmapped = config.model_copy(update={"id_column_no": 0})

        with pytest.raises(ValueError, match=r"'id'"):
            read_data_frame_from_file_path(file_path, unmapped, sync_context, Protocol())


@pytest.mark.unit
class TestBooleanUdfBlankPreservation:
    """An empty boolean cell is unset, not false.

    Coercing it to `falseValue` also re-fills an otherwise-empty row on the next
    write, which brings back the whole-file abort this feature removes.
    """

    @pytest.fixture
    def boolean_config(self, config: ExcelDefectClientConfig) -> ExcelDefectClientConfig:
        config.udfs = [
            UserDefiendAttributes(
                name="isOpen",
                column=3,
                type=ValueType.BOOLEAN,
                trueValue="1-yes",
                falseValue="2-no",
            )
        ]
        return config

    def test_read_mapping_keeps_blank_cells_blank(self, boolean_config: ExcelDefectClientConfig):
        df = pd.DataFrame({"isOpen": ["1-yes", "2-no", "", "   "]})

        map_boolean_values(boolean_config, df)

        assert df["isOpen"].tolist() == ["true", "false", "", ""]

    def test_write_mapping_keeps_blank_cells_blank(self, boolean_config: ExcelDefectClientConfig):
        df = pd.DataFrame({"isOpen": ["true", "false", ""]})

        _apply_boolean_udf_write_mapping(boolean_config, df, "isOpen")

        assert df["isOpen"].tolist() == ["1-yes", "2-no", ""]

    def test_blank_boolean_cell_round_trips_as_blank(self, boolean_config: ExcelDefectClientConfig):
        df = pd.DataFrame({"isOpen": ["1-yes", ""]})

        map_boolean_values(boolean_config, df)
        _apply_boolean_udf_write_mapping(boolean_config, df, "isOpen")

        assert df["isOpen"].tolist() == ["1-yes", ""]

    def test_empty_row_stays_blank_through_a_read_with_a_boolean_udf(
        self,
        tmp_path: Path,
        boolean_config: ExcelDefectClientConfig,
        sync_context: SyncContext,
    ):
        file_path = tmp_path / "defects.csv"
        file_path.write_text(
            "id,title,isOpen\nD-1,First,1-yes\n,,\nD-2,Second,2-no\n", encoding="utf-8"
        )

        df = read_data_frame_from_file_path(file_path, boolean_config, sync_context, Protocol())

        assert df["isOpen"].tolist() == ["true", "", "false"]
        assert is_blank_row(df.iloc[1]) is True

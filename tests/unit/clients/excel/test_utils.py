from collections.abc import Callable
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from unittest.mock import patch

import openpyxl
import pandas as pd
import pytest

from testbench_defect_service.clients.excel.config import (
    ControlFields,
    ExcelDefectClientConfig,
    Transition,
)
from testbench_defect_service.clients.excel.utils import (
    add_general_warning_once,
    check_defect_transitions,
    coerce_cell_to_string,
    create_defect_data_frame,
    get_column_mapping_for_config,
    get_visible_sheets,
    is_blank_cell,
    is_blank_row,
    map_and_rename_columns,
    optional_row_value,
    parse_boolean_udf_value,
    read_header_columns_from_file_path,
    resolve_sheet_name,
    resolve_visible_sheet_name,
    row_value,
    split_references,
    to_python_datetime_format,
    validate_control_fields,
)
from testbench_defect_service.models.defects import (
    Defect,
    Login,
    Protocol,
    ProtocolCode,
    SyncContext,
    UserDefinedFieldProperties,
)

# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

# Fields that are the same across every config in this test module.
_BASE_CONFIG_KWARGS: dict[str, Any] = {
    "system_name": "Excel",
    "simple_date_format": "yyyy-MM-dd",
    "references_separator": ";",
    "id_prefix": "D-",
    "defect_id_starting_value": "1",
    "defect_id_digit_numbers": 4,
}

# Default 1-based column positions used by most tests.
_DEFAULT_COLUMN_KWARGS: dict[str, Any] = {
    "id_column_no": 1,
    "title_column_no": 2,
    "references_column_no": 3,
    "discoverer_column_no": 4,
    "lastedit_column_no": 5,
    "description_column_no": 6,
}


def _make_csv_config(
    tmp_path: Path, header_line: int = 1, separator: str = ","
) -> ExcelDefectClientConfig:
    return ExcelDefectClientConfig(
        **_BASE_CONFIG_KWARGS,
        **_DEFAULT_COLUMN_KWARGS,
        excel_file_path=tmp_path,
        file_type=".csv",
        separator=separator,
        defects_data_header_line=header_line,
        defects_data_starting_line=header_line + 1,
    )


def _make_xlsx_config(tmp_path: Path) -> ExcelDefectClientConfig:
    return ExcelDefectClientConfig(
        **_BASE_CONFIG_KWARGS,
        **_DEFAULT_COLUMN_KWARGS,
        excel_file_path=tmp_path,
        file_type=".xlsx",
        separator=",",
        defects_data_header_line=1,
        defects_data_starting_line=2,
    )


@pytest.fixture
def column_mapping_config() -> Callable[..., ExcelDefectClientConfig]:
    def _factory(  # noqa: PLR0913
        control_fields: list[ControlFields] | None = None,
        *,
        id_column_no: int = 1,
        title_column_no: int = 2,
        references_column_no: int = 3,
        discoverer_column_no: int = 4,
        lastedit_column_no: int = 5,
        description_column_no: int = 6,
    ) -> ExcelDefectClientConfig:
        return ExcelDefectClientConfig(
            **_BASE_CONFIG_KWARGS,
            excel_file_path=Path("/tmp"),
            file_type=".csv",
            separator=",",
            defects_data_header_line=1,
            defects_data_starting_line=2,
            id_column_no=id_column_no,
            title_column_no=title_column_no,
            references_column_no=references_column_no,
            discoverer_column_no=discoverer_column_no,
            lastedit_column_no=lastedit_column_no,
            description_column_no=description_column_no,
            control_fields=control_fields or [],
        )

    return _factory


@pytest.fixture
def sync_context() -> Callable[..., SyncContext]:
    def _factory(
        status: str | None = None,
        priority: str | None = None,
        classification: str | None = None,
    ) -> SyncContext:
        return SyncContext(
            statusAttribute=status,
            priorityAttribute=priority,
            classAttribute=classification,
        )

    return _factory


# ---------------------------------------------------------------------------
# Tests: read_header_columns_from_file_path
# ---------------------------------------------------------------------------


@pytest.mark.unit
class TestReadHeaderColumnsFromFilePath:
    def test_csv_returns_1_indexed_dict(self, tmp_path: Path) -> None:
        csv_file = tmp_path / "defects.csv"
        csv_file.write_text("id,title,status\nD-1,Bug,Open\n", encoding="utf-8")

        result = read_header_columns_from_file_path(csv_file, _make_csv_config(tmp_path))

        assert result == {1: "id", 2: "title", 3: "status"}

    def test_tsv_returns_correct_columns(self, tmp_path: Path) -> None:
        tsv_file = tmp_path / "defects.tsv"
        tsv_file.write_text("id\ttitle\tstatus\nD-1\tBug\tOpen\n", encoding="utf-8")

        result = read_header_columns_from_file_path(tsv_file, _make_csv_config(tmp_path))

        assert result == {1: "id", 2: "title", 3: "status"}

    def test_non_default_header_line(self, tmp_path: Path) -> None:
        csv_file = tmp_path / "defects.csv"
        csv_file.write_text(
            "metadata\nmore metadata\nid,title,status\nD-1,Bug,Open\n",
            encoding="utf-8",
        )

        result = read_header_columns_from_file_path(
            csv_file, _make_csv_config(tmp_path, header_line=3)
        )

        assert result == {1: "id", 2: "title", 3: "status"}

    def test_trailing_empty_columns_are_stripped(self, tmp_path: Path) -> None:
        csv_file = tmp_path / "defects.csv"
        csv_file.write_text("id,title,,\nD-1,Bug,,\n", encoding="utf-8")

        result = read_header_columns_from_file_path(csv_file, _make_csv_config(tmp_path))

        assert result == {1: "id", 2: "title"}

    def test_xlsx_returns_correct_columns(self, tmp_path: Path) -> None:
        xlsx_file = tmp_path / "defects.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Sheet1"
        ws.append(["id", "title", "status"])
        ws.append(["D-1", "Bug", "Open"])
        wb.save(xlsx_file)

        result = read_header_columns_from_file_path(xlsx_file, _make_xlsx_config(tmp_path))

        assert result == {1: "id", 2: "title", 3: "status"}

    def test_unsupported_extension_raises(self, tmp_path: Path) -> None:
        unsupported_file = tmp_path / "defects.ods"
        unsupported_file.write_text("id,title\n", encoding="utf-8")

        with pytest.raises(ValueError, match="Unsupported file format"):
            read_header_columns_from_file_path(unsupported_file, _make_csv_config(tmp_path))

    def test_header_line_beyond_file_raises(self, tmp_path: Path) -> None:
        csv_file = tmp_path / "defects.csv"
        csv_file.write_text("id,title\n", encoding="utf-8")

        with pytest.raises(ValueError, match="Header row"):
            read_header_columns_from_file_path(csv_file, _make_csv_config(tmp_path, header_line=5))

    def test_utf8_bom_is_stripped_from_first_header(self, tmp_path: Path) -> None:
        csv_file = tmp_path / "defects.csv"
        csv_file.write_text("id,title,status\nD-1,Bug,Open\n", encoding="utf-8-sig")

        result = read_header_columns_from_file_path(csv_file, _make_csv_config(tmp_path))

        assert result == {1: "id", 2: "title", 3: "status"}

    def test_protocol_none_does_not_raise(self, tmp_path: Path) -> None:
        csv_file = tmp_path / "defects.csv"
        csv_file.write_text("id,title\nD-1,Bug\n", encoding="utf-8")

        result = read_header_columns_from_file_path(
            csv_file, _make_csv_config(tmp_path), protocol=None
        )

        assert result == {1: "id", 2: "title"}


# ---------------------------------------------------------------------------
# Tests: get_column_mapping_for_config
# ---------------------------------------------------------------------------


@pytest.mark.unit
class TestGetColumnMappingForConfig:
    def test_base_fields_no_control_attributes(self, column_mapping_config, sync_context) -> None:
        result = get_column_mapping_for_config(column_mapping_config(), sync_context())

        assert result is not None
        assert result[0] == ["id"]
        assert result[1] == ["title"]
        assert result[2] == ["references"]
        assert result[3] == ["reporter"]
        assert result[4] == ["lastEdited"]
        assert result[5] == ["description"]

    def test_includes_matched_control_fields(self, column_mapping_config, sync_context) -> None:
        control_fields = [
            ControlFields(name="status", column_number=7, values=["Open", "Closed"]),
            ControlFields(name="priority", column_number=8, values=["High", "Low"]),
            ControlFields(name="classification", column_number=9, values=["Bug"]),
        ]
        config = column_mapping_config(control_fields)
        ctx = sync_context(status="status", priority="priority", classification="classification")

        result = get_column_mapping_for_config(config, ctx)

        assert result is not None
        assert result[6] == ["status"]
        assert result[7] == ["priority"]
        assert result[8] == ["classification"]

    def test_missing_control_field_with_protocol_returns_none_and_logs_error(
        self, column_mapping_config, sync_context
    ) -> None:
        protocol = Protocol()

        result = get_column_mapping_for_config(
            column_mapping_config(), sync_context(status="status"), protocol
        )

        assert result is None
        assert protocol.generalErrors
        assert any(e.message is not None and "status" in e.message for e in protocol.generalErrors)

    def test_missing_control_field_without_protocol_raises(
        self, column_mapping_config, sync_context
    ) -> None:
        with pytest.raises(ValueError, match="status"):
            get_column_mapping_for_config(
                column_mapping_config(),
                sync_context(status="status", priority="priority"),
                protocol=None,
            )

    def test_multiple_fields_mapped_to_same_column(
        self, column_mapping_config, sync_context
    ) -> None:
        config = column_mapping_config(id_column_no=1, title_column_no=1)

        result = get_column_mapping_for_config(config, sync_context())

        assert result is not None
        assert "id" in result[0]
        assert "title" in result[0]

    def test_skips_zero_and_negative_column_numbers(
        self, column_mapping_config, sync_context
    ) -> None:
        config = column_mapping_config(id_column_no=0, title_column_no=-1)

        result = get_column_mapping_for_config(config, sync_context())

        assert result is not None
        assert not any("id" in names for names in result.values())
        assert not any("title" in names for names in result.values())

    def test_all_three_missing_control_fields_logs_all_errors(
        self, column_mapping_config, sync_context
    ) -> None:
        ctx = sync_context(status="status", priority="priority", classification="classification")
        protocol = Protocol()

        result = get_column_mapping_for_config(column_mapping_config(), ctx, protocol)

        assert result is None
        assert protocol.generalErrors is not None
        assert len(protocol.generalErrors) == 3


# ---------------------------------------------------------------------------
# Tests: map_and_rename_columns
# ---------------------------------------------------------------------------


@pytest.mark.unit
class TestMapAndRenameColumns:
    def test_renames_columns_to_header_values(self, column_mapping_config, sync_context) -> None:
        config = column_mapping_config()
        ctx = sync_context()
        header = {1: "Defect ID", 2: "Summary", 3: "Refs", 4: "Reporter", 5: "Last Edit", 6: "Desc"}
        df = pd.DataFrame(
            {
                "id": ["D-1"],
                "title": ["Bug"],
                "references": ["R-1"],
                "reporter": ["Alice"],
                "lastEdited": ["2024-01-01"],
                "description": ["A bug"],
            }
        )

        result = map_and_rename_columns(ctx, config, header, df)

        assert list(result.columns) == [
            "Defect ID",
            "Summary",
            "Refs",
            "Reporter",
            "Last Edit",
            "Desc",
        ]

    def test_df_columns_not_in_header_are_kept_unchanged(
        self, column_mapping_config, sync_context
    ) -> None:
        config = column_mapping_config()
        ctx = sync_context()
        header = {1: "Defect ID"}
        df = pd.DataFrame({"id": ["D-1"], "title": ["Bug"]})

        result = map_and_rename_columns(ctx, config, header, df)

        assert "Defect ID" in result.columns
        assert "title" in result.columns

    def test_extra_df_columns_outside_mapping_are_preserved(
        self, column_mapping_config, sync_context
    ) -> None:
        config = column_mapping_config()
        ctx = sync_context()
        header = {1: "Defect ID", 2: "Summary", 3: "Refs", 4: "Reporter", 5: "Last Edit", 6: "Desc"}
        df = pd.DataFrame(
            {
                "id": ["D-1"],
                "title": ["Bug"],
                "references": ["R-1"],
                "reporter": ["Alice"],
                "lastEdited": ["2024-01-01"],
                "description": ["A bug"],
                "extra": ["value"],
            }
        )

        result = map_and_rename_columns(ctx, config, header, df)

        assert "extra" in result.columns


# ---------------------------------------------------------------------------
# Tests: coerce_cell_to_string
# ---------------------------------------------------------------------------


@pytest.mark.unit
class TestCoerceCellToString:
    def test_none_returns_empty_string(self) -> None:
        assert coerce_cell_to_string(None) == ""

    def test_float_nan_returns_empty_string(self) -> None:
        assert coerce_cell_to_string(float("nan")) == ""

    def test_integer_float_returns_integer_string(self) -> None:
        assert coerce_cell_to_string(1.0) == "1"

    def test_non_integer_float_returns_decimal_string(self) -> None:
        assert coerce_cell_to_string(1.5) == "1.5"

    def test_string_is_stripped(self) -> None:
        assert coerce_cell_to_string("  hello  ") == "hello"

    def test_empty_string_stays_empty(self) -> None:
        assert coerce_cell_to_string("") == ""

    def test_integer_returns_string(self) -> None:
        assert coerce_cell_to_string(42) == "42"

    def test_boolean_returns_string(self) -> None:
        assert coerce_cell_to_string(True) == "True"


# ---------------------------------------------------------------------------
# Tests: get_visible_sheets
# ---------------------------------------------------------------------------


@pytest.mark.unit
class TestGetVisibleSheets:
    def test_xlsx_returns_visible_sheet_names(self, tmp_path: Path) -> None:
        xlsx_file = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        assert wb.active is not None
        wb.active.title = "Defects"
        wb.save(xlsx_file)

        result = get_visible_sheets(xlsx_file)

        assert result == ["Defects"]

    def test_xlsx_excludes_hidden_sheet(self, tmp_path: Path) -> None:
        xlsx_file = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        assert wb.active is not None
        wb.active.title = "Visible"
        hidden = wb.create_sheet("Hidden")
        hidden.sheet_state = "hidden"
        wb.save(xlsx_file)

        result = get_visible_sheets(xlsx_file)

        assert result == ["Visible"]

    def test_xlsx_multiple_visible_sheets(self, tmp_path: Path) -> None:
        xlsx_file = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        assert wb.active is not None
        wb.active.title = "Sheet1"
        wb.create_sheet("Sheet2")
        wb.save(xlsx_file)

        result = get_visible_sheets(xlsx_file)

        assert result == ["Sheet1", "Sheet2"]

    def test_unsupported_extension_raises(self, tmp_path: Path) -> None:
        other_file = tmp_path / "test.ods"
        other_file.write_text("", encoding="utf-8")

        with pytest.raises(ValueError, match="Unsupported Excel file format"):
            get_visible_sheets(other_file)

    def test_xls_file_with_xlsx_content_raises_helpful_error(self, tmp_path: Path) -> None:
        mislabeled_file = tmp_path / "test.xls"
        wb = openpyxl.Workbook()
        wb.save(mislabeled_file)

        with pytest.raises(ValueError, match="xlsx content"):
            get_visible_sheets(mislabeled_file)


# ---------------------------------------------------------------------------
# Tests: resolve_sheet_name
# ---------------------------------------------------------------------------


@pytest.mark.unit
class TestResolveSheetName:
    def test_empty_configured_name_returns_first_sheet(self) -> None:
        result = resolve_sheet_name("", ["Sheet1", "Sheet2"], "file.xlsx")

        assert result == "Sheet1"

    def test_none_configured_name_returns_first_sheet(self) -> None:
        result = resolve_sheet_name(None, ["Sheet1", "Sheet2"], "file.xlsx")

        assert result == "Sheet1"

    def test_configured_name_found_returns_it(self) -> None:
        result = resolve_sheet_name("Sheet2", ["Sheet1", "Sheet2"], "file.xlsx")

        assert result == "Sheet2"

    def test_configured_name_not_found_falls_back_to_first(self) -> None:
        result = resolve_sheet_name("Missing", ["Sheet1", "Sheet2"], "file.xlsx")

        assert result == "Sheet1"

    def test_configured_name_not_found_adds_warning_to_protocol(self) -> None:
        protocol = Protocol()

        result = resolve_sheet_name("Missing", ["Sheet1"], "file.xlsx", protocol)

        assert result == "Sheet1"
        assert protocol.generalWarnings
        assert any(
            w.message is not None and "Missing" in w.message for w in protocol.generalWarnings
        )


# ---------------------------------------------------------------------------
# Tests: resolve_visible_sheet_name
# ---------------------------------------------------------------------------


@pytest.mark.unit
class TestResolveVisibleSheetName:
    def test_no_configured_sheet_returns_first_visible(self, tmp_path: Path) -> None:
        xlsx_file = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        assert wb.active is not None
        wb.active.title = "FirstSheet"
        wb.save(xlsx_file)
        config = _make_xlsx_config(tmp_path)

        result = resolve_visible_sheet_name(xlsx_file, config)

        assert result == "FirstSheet"

    def test_configured_sheet_found_returns_it(self, tmp_path: Path) -> None:
        xlsx_file = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        assert wb.active is not None
        wb.active.title = "Default"
        wb.create_sheet("Defects")
        wb.save(xlsx_file)
        config = ExcelDefectClientConfig(
            **_BASE_CONFIG_KWARGS,
            **_DEFAULT_COLUMN_KWARGS,
            excel_file_path=tmp_path,
            file_type=".xlsx",
            separator=",",
            defects_data_header_line=1,
            defects_data_starting_line=2,
            worksheet_name="Defects",
        )

        result = resolve_visible_sheet_name(xlsx_file, config)

        assert result == "Defects"

    def test_configured_sheet_not_found_falls_back_to_first(self, tmp_path: Path) -> None:
        xlsx_file = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        assert wb.active is not None
        wb.active.title = "Sheet1"
        wb.save(xlsx_file)
        config = ExcelDefectClientConfig(
            **_BASE_CONFIG_KWARGS,
            **_DEFAULT_COLUMN_KWARGS,
            excel_file_path=tmp_path,
            file_type=".xlsx",
            separator=",",
            defects_data_header_line=1,
            defects_data_starting_line=2,
            worksheet_name="NonExistent",
        )

        result = resolve_visible_sheet_name(xlsx_file, config)

        assert result == "Sheet1"

    def test_no_visible_sheets_raises(self, tmp_path: Path) -> None:

        xlsx_file = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        assert wb.active is not None
        wb.active.title = "Sheet1"
        wb.save(xlsx_file)
        config = _make_xlsx_config(tmp_path)

        with (
            patch(
                "testbench_defect_service.clients.excel.utils.get_visible_sheets",
                return_value=[],
            ),
            pytest.raises(ValueError, match="No visible worksheets"),
        ):
            resolve_visible_sheet_name(xlsx_file, config)


# ---------------------------------------------------------------------------
# Tests: create_defect_data_frame
# ---------------------------------------------------------------------------


@pytest.mark.unit
class TestCreateDefectDataFrame:
    def test_defect_dataframe_creation_no_udfs_no_references(self, tmp_path: Path):
        csv_file = tmp_path / "defects.csv"
        time = datetime(2026, 6, 1, 9, 35, 7, 118178, tzinfo=timezone.utc)
        defect = Defect(
            title="title",
            description="description",
            reporter="reporter",
            status="status",
            classification="classification",
            priority="priority",
            lastEdited=time,
            references=[],
            userDefinedFields=[],
            principal=Login(username="username", password="password"),
        )

        df = create_defect_data_frame(
            defect=defect,
            config=_make_csv_config(csv_file),
            defect_id="Bug 007",
            protocol=Protocol(),
        )

        assert df.iloc[0]["title"] == defect.title
        assert df.iloc[0]["description"] == defect.description
        assert df.iloc[0]["reporter"] == defect.reporter
        assert df.iloc[0]["status"] == defect.status
        assert df.iloc[0]["classification"] == defect.classification
        assert df.iloc[0]["lastEdited"] == "2026-06-01"

    def test_defect_dataframe_creation_single_reference(self, tmp_path: Path):
        csv_file = tmp_path / "defects.csv"
        time = datetime(2026, 6, 1, 9, 35, 7, 118178, tzinfo=timezone.utc)
        defect = Defect(
            title="title",
            description="description",
            reporter="reporter",
            status="status",
            classification="classification",
            priority="priority",
            lastEdited=time,
            references=["reference_1"],
            userDefinedFields=[],
            principal=Login(username="username", password="password"),
        )

        df = create_defect_data_frame(
            defect=defect,
            config=_make_csv_config(csv_file),
            defect_id="Bug 007",
            protocol=Protocol(),
        )

        assert df.iloc[0]["references"] == "reference_1"

    def test_defect_dataframe_creation_with_multiple_references(self, tmp_path: Path):
        csv_file = tmp_path / "defects.csv"
        time = datetime(2026, 6, 1, 9, 35, 7, 118178, tzinfo=timezone.utc)
        defect = Defect(
            title="title",
            description="description",
            reporter="reporter",
            status="status",
            classification="classification",
            priority="priority",
            lastEdited=time,
            references=["reference_1", "reference_2"],
            userDefinedFields=[],
            principal=Login(username="username", password="password"),
        )

        df = create_defect_data_frame(
            defect=defect,
            config=_make_csv_config(csv_file),
            defect_id="Bug 007",
            protocol=Protocol(),
        )

        assert df.iloc[0]["references"] == "reference_1;reference_2"

    def test_defect_dataframe_creation_single_udf(self, tmp_path: Path):
        csv_file = tmp_path / "defects.csv"
        time = datetime(2026, 6, 1, 9, 35, 7, 118178, tzinfo=timezone.utc)
        defect = Defect(
            title="title",
            description="description",
            reporter="reporter",
            status="status",
            classification="classification",
            priority="priority",
            lastEdited=time,
            references=[],
            userDefinedFields=[UserDefinedFieldProperties(name="isOpen", value="true")],
            principal=Login(username="username", password="password"),
        )

        df = create_defect_data_frame(
            defect=defect,
            config=_make_csv_config(csv_file),
            defect_id="Bug 007",
            protocol=Protocol(),
        )

        assert df.iloc[0]["isOpen"] == "true"

    def test_defect_dataframe_with_multiple_udfs(self, tmp_path: Path):
        csv_file = tmp_path / "defects.csv"
        time = datetime(2026, 6, 1, 9, 35, 7, 118178, tzinfo=timezone.utc)
        defect = Defect(
            title="title",
            description="description",
            reporter="reporter",
            status="status",
            classification="classification",
            priority="priority",
            lastEdited=time,
            references=[],
            userDefinedFields=[
                UserDefinedFieldProperties(name="isOpen", value="true"),
                UserDefinedFieldProperties(name="attributes", value="new"),
            ],
            principal=Login(username="username", password="password"),
        )

        df = create_defect_data_frame(
            defect=defect,
            config=_make_csv_config(csv_file),
            defect_id="Bug 007",
            protocol=Protocol(),
        )

        assert df.iloc[0]["isOpen"] == "true"
        assert df.iloc[0]["attributes"] == "new"


# ---------------------------------------------------------------------------
# Tests: to_python_datetime_format
# ---------------------------------------------------------------------------


@pytest.mark.unit
class TestToPythonDatetimeFormat:
    def test_to_python_datetime_format(self):
        simple_date_format = "yyyy-yy-dd-HH-hh-ss-MM"
        python_fromat = to_python_datetime_format(simple_date_format)

        assert python_fromat == "%Y-%y-%d-%H-%I-%S-%m"

    def test_to_python_datetime_format_with_none(self):
        assert to_python_datetime_format(None) is None

    def test_to_python_datetime_format_mm_to_upper_m(self):
        simple_date_format = "HH-mm"
        python_fromat = to_python_datetime_format(simple_date_format)

        simple_date_format_hh = "hh-mm"
        python_fromat_hh = to_python_datetime_format(simple_date_format_hh)

        assert python_fromat == "%H-%M"
        assert python_fromat_hh == "%I-%M"

    def test_to_python_datetime_format_mm_to_m(self):
        simple_date_format = "mm"
        python_fromat = to_python_datetime_format(simple_date_format)

        assert python_fromat == "%m"


# ---------------------------------------------------------------------------
# Tests: row_value
# ---------------------------------------------------------------------------


@pytest.mark.unit
class TestRowValue:
    def test_row_value(self):
        row = pd.Series(data={"title": "title_1", "status": "status_1"}, index=["title", "status"])
        value = row_value(row, "title")

        assert value == "title_1"

    def test_row_value_when_field_not_exists(self):
        row = pd.Series(data={"title": "title_1", "status": "status_1"}, index=["title", "status"])
        value = row_value(row, "description")

        assert value == ""


# ---------------------------------------------------------------------------
# Tests: optional_row_value
# ---------------------------------------------------------------------------


@pytest.mark.unit
class TestOptionalRowValue:
    def test_optional_row_value(self):
        row = pd.Series(data={"title": "title_1", "status": "status_1"}, index=["title", "status"])
        value = optional_row_value(row, "title")

        assert value == "title_1"


# ---------------------------------------------------------------------------
# Tests: is_blank_cell
# ---------------------------------------------------------------------------


@pytest.mark.unit
class TestIsBlankCell:
    def test_empty_string_is_blank(self):
        assert is_blank_cell("") is True

    def test_whitespace_is_blank(self):
        assert is_blank_cell("   ") is True

    def test_none_is_blank(self):
        assert is_blank_cell(None) is True

    def test_nan_is_blank(self):
        assert is_blank_cell(float("nan")) is True

    def test_value_is_not_blank(self):
        assert is_blank_cell("D-1") is False

    def test_zero_is_not_blank(self):
        assert is_blank_cell(0) is False


# ---------------------------------------------------------------------------
# Tests: is_blank_row
# ---------------------------------------------------------------------------


@pytest.mark.unit
class TestIsBlankRow:
    def test_all_cells_empty(self):
        row = pd.Series({"id": "", "title": "", "status": ""})
        assert is_blank_row(row) is True

    def test_whitespace_only_cells(self):
        row = pd.Series({"id": " ", "title": "\t", "status": ""})
        assert is_blank_row(row) is True

    def test_missing_and_none_cells(self):
        row = pd.Series({"id": None, "title": float("nan"), "status": ""})
        assert is_blank_row(row) is True

    def test_one_populated_cell_is_not_blank(self):
        row = pd.Series({"id": "", "title": "Only a title", "status": ""})
        assert is_blank_row(row) is False

    def test_populated_id_is_not_blank(self):
        row = pd.Series({"id": "D-1", "title": "", "status": ""})
        assert is_blank_row(row) is False


# ---------------------------------------------------------------------------
# Tests: split_references
# ---------------------------------------------------------------------------


@pytest.mark.unit
class TestSplitReferences:
    def test_split_references(self, tmp_path: Path):
        csv_file = tmp_path / "defects.csv"
        value = "Hello;Wor;ld"

        result = split_references(value, _make_xlsx_config(csv_file))

        assert result == ["Hello", "Wor", "ld"]

    def test_split_references_single_reference(self, tmp_path: Path):
        csv_file = tmp_path / "defects.csv"
        value = "Hello"

        result = split_references(value, _make_xlsx_config(csv_file))

        assert result == ["Hello"]

    def test_split_references_with_empty_value(self, tmp_path: Path):
        csv_file = tmp_path / "defects.csv"
        value = ""

        result = split_references(value, _make_xlsx_config(csv_file))

        assert result == []


# ---------------------------------------------------------------------------
# Tests: parse_boolean_udf_value
# ---------------------------------------------------------------------------


@pytest.mark.unit
class TestParseBooleanUdfValue:
    def test_returns_true_for_matching_true_value(self):
        assert parse_boolean_udf_value("yes", "yes", "no") is True

    def test_returns_false_for_matching_false_value(self):
        assert parse_boolean_udf_value("no", "yes", "no") is False

    def test_returns_none_for_unrecognized_value(self):
        assert parse_boolean_udf_value("maybe", "yes", "no") is None

    def test_matching_is_case_insensitive(self):
        assert parse_boolean_udf_value("YES", "yes", "no") is True
        assert parse_boolean_udf_value("NO", "yes", "no") is False

    def test_strips_whitespace_from_raw_value(self):
        assert parse_boolean_udf_value("  yes  ", "yes", "no") is True

    def test_strips_whitespace_from_configured_values(self):
        assert parse_boolean_udf_value("yes", "  yes  ", "  no  ") is True
        assert parse_boolean_udf_value("no", "  yes  ", "  no  ") is False

    def test_defaults_to_true_when_true_value_is_none(self):
        assert parse_boolean_udf_value("true", None, "false") is True

    def test_defaults_to_false_when_false_value_is_none(self):
        assert parse_boolean_udf_value("false", "true", None) is False

    def test_defaults_both_when_both_values_are_none(self):
        assert parse_boolean_udf_value("true", None, None) is True
        assert parse_boolean_udf_value("false", None, None) is False
        assert parse_boolean_udf_value("yes", None, None) is None

    def test_empty_string_raw_value_returns_none(self):
        assert parse_boolean_udf_value("", "yes", "no") is None

    def test_empty_string_true_value_falls_back_to_default(self):
        assert parse_boolean_udf_value("true", "", "no") is True
        assert parse_boolean_udf_value("", "", "no") is None


# ---------------------------------------------------------------------------
# Tests: parse_boolean_udf_value
# ---------------------------------------------------------------------------


@pytest.mark.unit
class TestAddGeneralWarningOnce:
    def test_adds_warning_when_protocol_is_empty(self):
        protocol = Protocol()
        add_general_warning_once(protocol, "some warning", ProtocolCode.IMPORT_WARNING)
        assert len(protocol.generalWarnings) == 1
        assert protocol.generalWarnings[0].message == "some warning"
        assert protocol.generalWarnings[0].code == ProtocolCode.IMPORT_WARNING

    def test_does_not_add_duplicate_warning(self):
        protocol = Protocol()
        add_general_warning_once(protocol, "some warning", ProtocolCode.IMPORT_WARNING)
        add_general_warning_once(protocol, "some warning", ProtocolCode.IMPORT_WARNING)
        assert len(protocol.generalWarnings) == 1

    def test_adds_different_message_as_new_warning(self):
        protocol = Protocol()
        add_general_warning_once(protocol, "warning A", ProtocolCode.IMPORT_WARNING)
        add_general_warning_once(protocol, "warning B", ProtocolCode.IMPORT_WARNING)
        assert len(protocol.generalWarnings) == 2

    def test_adds_same_message_with_different_code_as_new_warning(self):
        protocol = Protocol()
        add_general_warning_once(protocol, "some warning", ProtocolCode.IMPORT_WARNING)
        add_general_warning_once(protocol, "some warning", ProtocolCode.INSERT_WARNING)
        assert len(protocol.generalWarnings) == 2

    def test_does_not_add_duplicate_across_multiple_calls(self):
        protocol = Protocol()
        for _ in range(5):
            add_general_warning_once(protocol, "repeated", ProtocolCode.IMPORT_WARNING)
        assert len(protocol.generalWarnings) == 1

    def test_works_when_general_warnings_is_none(self):
        protocol = Protocol(generalWarnings=None)
        add_general_warning_once(protocol, "some warning", ProtocolCode.IMPORT_WARNING)
        assert protocol.generalWarnings is not None
        assert len(protocol.generalWarnings) == 1


# ---------------------------------------------------------------------------
# Tests: validate_control_fields
# ---------------------------------------------------------------------------


_DUMMY_LOGIN = Login(username="", password="")
_DUMMY_LAST_EDITED = datetime(2024, 1, 1, tzinfo=timezone.utc)


def _make_defect(
    status: str = "New",
    priority: str = "High",
    classification: str = "Bug",
) -> Defect:
    return Defect(
        status=status,
        priority=priority,
        classification=classification,
        lastEdited=_DUMMY_LAST_EDITED,
        principal=_DUMMY_LOGIN,
    )


def _make_config(control_fields: list[ControlFields]) -> ExcelDefectClientConfig:
    return ExcelDefectClientConfig(
        **_BASE_CONFIG_KWARGS,
        **_DEFAULT_COLUMN_KWARGS,
        excel_file_path=Path("/tmp"),
        file_type=".xlsx",
        separator=",",
        defects_data_header_line=1,
        defects_data_starting_line=2,
        control_fields=control_fields,
    )


def _make_sync_context(
    status: str | None = "status",
    priority: str | None = "priority",
    classification: str | None = "classification",
) -> SyncContext:
    return SyncContext(
        statusAttribute=status,
        priorityAttribute=priority,
        classAttribute=classification,
    )


@pytest.mark.unit
class TestValidateControlFields:
    def test_all_valid_returns_true(self) -> None:
        defect = _make_defect(status="New", priority="High", classification="Bug")
        config = _make_config(
            [
                ControlFields(name="status", column_number=1, values=["New", "Done"]),
                ControlFields(name="priority", column_number=2, values=["High", "Low"]),
                ControlFields(name="classification", column_number=3, values=["Bug", "Error"]),
            ]
        )
        sync_context = _make_sync_context()
        protocol = Protocol()

        result = validate_control_fields(defect, config, sync_context, protocol)

        assert result is True
        assert not protocol.errors

    def test_invalid_value_returns_false_and_adds_error(self) -> None:
        defect = _make_defect(status="Invalid")
        config = _make_config(
            [
                ControlFields(name="status", column_number=1, values=["New", "Done"]),
                ControlFields(name="priority", column_number=2, values=["High", "Low"]),
                ControlFields(name="classification", column_number=3, values=["Bug", "Error"]),
            ]
        )
        sync_context = _make_sync_context()
        protocol = Protocol()

        result = validate_control_fields(defect, config, sync_context, protocol)

        assert result is False
        assert protocol.errors is not None
        assert "status" in protocol.errors
        error = protocol.errors["status"][0]
        assert error.message is not None
        assert "Invalid" in error.message
        assert error.code == ProtocolCode.PUBLISH_ERROR

    def test_missing_control_field_for_required_attribute_returns_false(self) -> None:
        defect = _make_defect(status="New", priority="High", classification="Bug")
        # Only status and priority configured — classification is missing
        config = _make_config(
            [
                ControlFields(name="status", column_number=1, values=["New", "Done"]),
                ControlFields(name="priority", column_number=2, values=["High", "Low"]),
            ]
        )
        sync_context = _make_sync_context()
        protocol = Protocol()

        result = validate_control_fields(defect, config, sync_context, protocol)

        assert result is False
        # The reason must be recorded: create_defect aborts on a False return before anything
        # else inspects the column mapping, so this is the only chance to report it.
        assert protocol.errors is not None
        assert "classification" in protocol.errors
        error = protocol.errors["classification"][0]
        assert error.message is not None
        assert "not configured as a control field" in error.message
        assert error.code == ProtocolCode.PUBLISH_ERROR

    def test_no_required_attributes_returns_true(self) -> None:
        defect = _make_defect()
        config = _make_config([])
        sync_context = _make_sync_context(status=None, priority=None, classification=None)
        protocol = Protocol()

        result = validate_control_fields(defect, config, sync_context, protocol)

        assert result is True

    def test_no_control_fields_with_required_attributes_returns_false(self) -> None:
        defect = _make_defect()
        config = _make_config([])
        sync_context = _make_sync_context(status="status", priority=None, classification=None)
        protocol = Protocol()

        result = validate_control_fields(defect, config, sync_context, protocol)

        assert result is False

    def test_irrelevant_control_field_is_skipped(self) -> None:
        defect = _make_defect(status="New")
        config = _make_config(
            [
                ControlFields(name="status", column_number=1, values=["New", "Done"]),
                ControlFields(name="unrelated_field", column_number=9, values=["x", "y"]),
            ]
        )
        sync_context = _make_sync_context(status="status", priority=None, classification=None)
        protocol = Protocol()

        result = validate_control_fields(defect, config, sync_context, protocol)

        assert result is True

    def test_only_some_sync_attributes_set(self) -> None:
        defect = _make_defect(status="New", priority="High")
        config = _make_config(
            [
                ControlFields(name="status", column_number=1, values=["New", "Done"]),
                ControlFields(name="priority", column_number=2, values=["High", "Low"]),
            ]
        )
        sync_context = _make_sync_context(status="status", priority="priority", classification=None)
        protocol = Protocol()

        result = validate_control_fields(defect, config, sync_context, protocol)

        assert result is True


# ---------------------------------------------------------------------------
# Tests: check_defect_transitions
# ---------------------------------------------------------------------------


@pytest.mark.unit
class TestCheckDefectTransitions:
    def _make_df(self, current_status: str = "New") -> pd.DataFrame:
        return pd.DataFrame({"status": [current_status]})

    def _make_config_with_transitions(
        self, transitions: list[Transition]
    ) -> ExcelDefectClientConfig:
        return ExcelDefectClientConfig(
            **_BASE_CONFIG_KWARGS,
            **_DEFAULT_COLUMN_KWARGS,
            excel_file_path=Path("/tmp"),
            file_type=".xlsx",
            separator=",",
            defects_data_header_line=1,
            defects_data_starting_line=2,
            transitions=transitions,
        )

    def test_no_transitions_configured_returns_true(self) -> None:
        defect = _make_defect(status="Done")
        config = self._make_config_with_transitions([])
        df = self._make_df(current_status="New")
        protocol = Protocol()

        result = check_defect_transitions(defect, df, config, protocol)

        assert result is True
        assert not protocol.warnings

    def test_valid_transition_returns_true(self) -> None:
        defect = _make_defect(status="Done")
        config = self._make_config_with_transitions(
            [
                Transition(from_state="New", to_state="Done"),
            ]
        )
        df = self._make_df(current_status="New")
        protocol = Protocol()

        result = check_defect_transitions(defect, df, config, protocol)

        assert result is True
        assert not protocol.warnings

    def test_invalid_transition_returns_false_and_adds_warning(self) -> None:
        defect = _make_defect(status="Done")
        config = self._make_config_with_transitions(
            [
                Transition(from_state="New", to_state="InProgress"),
            ]
        )
        df = self._make_df(current_status="New")
        protocol = Protocol()

        result = check_defect_transitions(defect, df, config, protocol)

        assert result is False
        assert protocol.warnings is not None
        assert "Done" in protocol.warnings
        warning = protocol.warnings["Done"][0]
        assert warning.message is not None
        assert "New" in warning.message
        assert "Done" in warning.message
        assert warning.code == ProtocolCode.PUBLISH_ERROR

    def test_matching_transition_among_multiple_returns_true(self) -> None:
        defect = _make_defect(status="Done")
        config = self._make_config_with_transitions(
            [
                Transition(from_state="New", to_state="InProgress"),
                Transition(from_state="New", to_state="Done"),
                Transition(from_state="InProgress", to_state="Done"),
            ]
        )
        df = self._make_df(current_status="New")
        protocol = Protocol()

        result = check_defect_transitions(defect, df, config, protocol)

        assert result is True
        assert not protocol.warnings

    def test_no_matching_from_state_returns_false(self) -> None:
        defect = _make_defect(status="Done")
        config = self._make_config_with_transitions(
            [
                Transition(from_state="InProgress", to_state="Done"),
            ]
        )
        df = self._make_df(current_status="New")
        protocol = Protocol()

        result = check_defect_transitions(defect, df, config, protocol)

        assert result is False
        assert protocol.warnings is not None
        assert "Done" in protocol.warnings

    def test_same_status_transition_allowed_when_configured(self) -> None:
        defect = _make_defect(status="New")
        config = self._make_config_with_transitions(
            [
                Transition(from_state="New", to_state="New"),
            ]
        )
        df = self._make_df(current_status="New")
        protocol = Protocol()

        result = check_defect_transitions(defect, df, config, protocol)

        assert result is True

    def test_valid_transition_allows_same_status(self) -> None:
        defect = _make_defect(status="Done")
        config = self._make_config_with_transitions(
            [
                Transition(from_state="New", to_state="InProgress"),
                Transition(from_state="New", to_state="Done"),
                Transition(from_state="InProgress", to_state="Done"),
            ]
        )
        df = self._make_df(current_status="Done")
        protocol = Protocol()

        result = check_defect_transitions(defect, df, config, protocol)

        assert result is True
        assert not protocol.warnings

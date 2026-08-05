import csv
import math
from datetime import datetime
from pathlib import Path
from typing import Any, NoReturn

import openpyxl
import pandas as pd
import xlrd

from testbench_defect_service.clients.excel.config import ExcelDefectClientConfig
from testbench_defect_service.log import logger
from testbench_defect_service.models.defects import (
    Defect,
    Protocol,
    ProtocolCode,
    SyncContext,
)


def describe_write_error(exc: OSError | ValueError, file_path: Path) -> str:
    """Turn a write failure into a message that says what to do about it.

    A file that is open in Excel is readable but denies writers, so the raw
    '[Errno 13] Permission denied' surfaces without hinting at the cause.
    """
    if isinstance(exc, PermissionError):
        return (
            f"Cannot write to '{file_path.name}': the file is open in another program "
            "(for example Excel) or is write-protected. Please close the file and run the "
            "synchronization again."
        )
    return str(exc)


def read_header_columns_from_file_path(
    file_path: Path,
    config: ExcelDefectClientConfig,
    protocol: Protocol | None = None,
) -> dict[int, str]:
    header_values = _load_header_values(file_path, config, protocol)
    return dict(enumerate(header_values, start=1))


def get_column_mapping_for_config(
    config: ExcelDefectClientConfig, sync_context: SyncContext, protocol: Protocol | None = None
) -> dict[int, list[str]] | None:
    column_mapping: dict[int, list[str]] = {}

    status_column_no = None
    priority_column_no = None
    class_column_no = None

    for control_field in config.control_fields:
        if sync_context.statusAttribute == control_field.name:
            status_column_no = control_field.column_number
        if sync_context.priorityAttribute == control_field.name:
            priority_column_no = control_field.column_number
        if sync_context.classAttribute == control_field.name:
            class_column_no = control_field.column_number

    if _add_missing_control_field_errors(
        sync_context,
        protocol,
        status_column_no=status_column_no,
        priority_column_no=priority_column_no,
        class_column_no=class_column_no,
    ):
        return None

    base_columns = {
        "id": config.id_column_no,
        "title": config.title_column_no,
        "references": config.references_column_no,
        "reporter": config.discoverer_column_no,
        "lastEdited": config.lastedit_column_no,
        "description": config.description_column_no,
        "status": status_column_no,
        "classification": class_column_no,
        "priority": priority_column_no,
    }
    for field_name, column_number in base_columns.items():
        _register_column(column_mapping, column_number, field_name)

    for udf_config in config.udfs:
        _register_column(column_mapping, udf_config.column, udf_config.name)

    return column_mapping


def _add_missing_control_field_errors(
    sync_context: SyncContext,
    protocol: Protocol | None,
    *,
    status_column_no: int | None,
    priority_column_no: int | None,
    class_column_no: int | None,
) -> bool:
    messages: list[str] = []
    required_control_fields = (
        ("status", sync_context.statusAttribute, status_column_no),
        ("priority", sync_context.priorityAttribute, priority_column_no),
        ("classification", sync_context.classAttribute, class_column_no),
    )

    for defect_field, attribute_name, column_number in required_control_fields:
        if not attribute_name or column_number:
            continue
        messages.append(
            f"Cannot import Excel defects: sync attribute '{attribute_name}' for "
            f"'{defect_field}' is not configured in the Excel control fields."
        )

    if not messages:
        return False

    if protocol is None:
        raise ValueError("\n".join(messages))

    for message in messages:
        logger.warning(message)
        protocol.add_general_error(
            message,
            protocol_code=ProtocolCode.IMPORT_ERROR,
        )
    return True


def _register_column(
    column_mapping: dict[int, list[str]],
    column_number: int | None,
    field_name: str,
) -> None:
    if column_number is None or column_number <= 0:
        return
    column_idx = column_number - 1
    column_mapping.setdefault(column_idx, [])
    if field_name not in column_mapping[column_idx]:
        column_mapping[column_idx].append(field_name)


def map_and_rename_columns(
    sync_context: SyncContext,
    config: ExcelDefectClientConfig,
    header: dict[int, str],
    df: pd.DataFrame,
) -> pd.DataFrame:
    column_mapping = get_column_mapping_for_config(config, sync_context)
    if column_mapping is not None:
        rename_map = {
            logical_names[0]: header[idx + 1]
            for idx, logical_names in column_mapping.items()
            if logical_names and idx + 1 in header
        }
        df = df.rename(columns=rename_map)
    return df


def _load_header_values(
    file_path: Path,
    config: ExcelDefectClientConfig,
    protocol: Protocol | None = None,
) -> list[str]:
    suffix = file_path.suffix.lower()

    if suffix == ".xlsx":
        return _load_xlsx_header_values(file_path, config, protocol)

    if suffix == ".xls":
        return _load_xls_header_values(file_path, config, protocol)

    if suffix in (".csv", ".tsv", ".txt"):
        return _load_delimited_header_values(file_path, config)

    raise ValueError(f"Unsupported file format: {file_path.suffix}")


def _load_xlsx_header_values(
    file_path: Path,
    config: ExcelDefectClientConfig,
    protocol: Protocol | None = None,
) -> list[str]:
    header_row_idx = max(config.defects_data_header_line - 1, 0)
    sheet_name = resolve_visible_sheet_name(file_path, config, protocol)
    try:
        workbook = openpyxl.load_workbook(
            file_path,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except KeyError as e:
        raise ValueError(
            f"File '{file_path}' does not appear to be a valid xlsx file (it may be corrupted or in"
            "a different format): {e}"
        ) from e
    try:
        worksheet = workbook[sheet_name]
        header_row = next(
            worksheet.iter_rows(
                min_row=header_row_idx + 1,
                max_row=header_row_idx + 1,
                values_only=True,
            ),
            None,
        )
    finally:
        workbook.close()

    if header_row is None:
        _raise_missing_header_row(file_path, config.defects_data_header_line)
    return _normalize_header_values(header_row)


def _load_xls_header_values(
    file_path: Path,
    config: ExcelDefectClientConfig,
    protocol: Protocol | None = None,
) -> list[str]:
    header_row_idx = max(config.defects_data_header_line - 1, 0)
    sheet_name = resolve_visible_sheet_name(file_path, config, protocol)
    workbook = xlrd.open_workbook(str(file_path), on_demand=True)
    try:
        worksheet = workbook.sheet_by_name(sheet_name)
        if header_row_idx >= worksheet.nrows:
            _raise_missing_header_row(file_path, config.defects_data_header_line)
        return _normalize_header_values(list(worksheet.row_values(header_row_idx)))
    finally:
        workbook.release_resources()


def resolve_delimited_separator(file_path: Path, config: ExcelDefectClientConfig) -> str:
    if file_path.suffix.lower() == ".tsv":
        return "\t"
    separator = config.separator or ","
    if len(separator) != 1:
        raise ValueError(f"Unsupported separator '{separator}' for '{file_path.name}'.")
    return separator


def _load_delimited_header_values(
    file_path: Path,
    config: ExcelDefectClientConfig,
) -> list[str]:
    separator = resolve_delimited_separator(file_path, config)

    last_error: UnicodeDecodeError | None = None
    for encoding in ("utf-8-sig", "windows-1252"):
        try:
            return _read_delimited_header_values(
                file_path,
                separator,
                encoding,
                config.defects_data_header_line,
            )
        except UnicodeDecodeError as exc:
            last_error = exc

    if last_error is not None:
        raise last_error

    _raise_missing_header_row(file_path, config.defects_data_header_line)
    return []  # type: ignore[unreachable]


def _read_delimited_header_values(
    file_path: Path,
    separator: str,
    encoding: str,
    header_line: int,
) -> list[str]:
    header_row_idx = max(header_line - 1, 0)
    with file_path.open("r", encoding=encoding, newline="") as handle:
        reader = csv.reader(handle, delimiter=separator)
        for row_index, row in enumerate(reader):
            if row_index == header_row_idx:
                return _normalize_header_values(row)
    _raise_missing_header_row(file_path, header_line)
    return []  # type: ignore[unreachable]


def resolve_visible_sheet_name(
    file_path: Path,
    config: ExcelDefectClientConfig,
    protocol: Protocol | None = None,
) -> str:
    visible_sheets = get_visible_sheets(file_path)
    if not visible_sheets:
        raise ValueError(f"No visible worksheets found in '{file_path.name}'.")
    return resolve_sheet_name(
        (config.worksheet_name or ""),
        visible_sheets,
        file_path.name,
        protocol,
    )


def _raise_missing_header_row(file_path: Path, header_line: int) -> NoReturn:
    raise ValueError(f"Header row {header_line} not found in '{file_path.name}'.")


def _normalize_header_values(values: list[Any] | tuple[Any, ...]) -> list[str]:
    header_values = [coerce_cell_to_string(value) for value in values]
    while header_values and header_values[-1] == "":
        header_values.pop()
    return header_values


def coerce_cell_to_string(value: Any, date_format: str | None = None) -> str:
    if value is None or value is pd.NaT:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    if isinstance(value, datetime) and date_format:
        return value.strftime(date_format)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def get_visible_sheets(file_path: Path) -> list[str]:
    suffix = file_path.suffix.lower()
    if suffix == ".xlsx":
        try:
            xlsx_workbook = openpyxl.load_workbook(
                file_path,
                read_only=True,
                data_only=True,
                keep_links=False,
            )
        except KeyError as e:
            raise ValueError(
                f"File '{file_path}' does not appear to be a valid xlsx file (it may be corrupted"
                f"or in a different format): {e}"
            ) from e
        try:
            return [
                sheet.title for sheet in xlsx_workbook.worksheets if sheet.sheet_state == "visible"
            ]
        finally:
            xlsx_workbook.close()

    if suffix == ".xls":
        with file_path.open("rb") as handle:
            magic_bytes = handle.read(4)
        if magic_bytes == b"PK\x03\x04":
            raise ValueError(
                f"File '{file_path.name}' has a .xls extension but contains xlsx content. "
                "Rename the file to .xlsx or convert it to the legacy .xls format."
            )
        xls_workbook = xlrd.open_workbook(str(file_path), on_demand=True)
        try:
            return [sheet.name for sheet in xls_workbook.sheets() if sheet.visibility == 0]
        finally:
            xls_workbook.release_resources()

    raise ValueError(f"Unsupported Excel file format: '{suffix}'.")


def resolve_sheet_name(
    configured_sheet_name: str | None,
    visible_sheets: list[str],
    file_name: str,
    protocol: Protocol | None = None,
) -> str:
    if not configured_sheet_name:
        return visible_sheets[0]
    if configured_sheet_name in visible_sheets:
        return configured_sheet_name

    warning_message = (
        f"Worksheet '{configured_sheet_name}' was not found or is hidden in '{file_name}'. "
        f"Falling back to '{visible_sheets[0]}'."
    )
    logger.warning(warning_message)
    if protocol is not None:
        protocol.add_general_warning(
            warning_message,
            protocol_code=ProtocolCode.IMPORT_WARNING,
        )
    return visible_sheets[0]


def add_defect_to_dataframe(
    defect: Defect,
    config: ExcelDefectClientConfig,
    df: pd.DataFrame,
    protocol: Protocol,
) -> pd.DataFrame:
    prefix = config.id_prefix
    numeric_ids = [
        int(id_val.replace(prefix, ""))
        for id_val in df["id"]
        if id_val.startswith(prefix) and id_val[len(prefix) :].isdigit()
    ]
    max_int = (max(numeric_ids) if numeric_ids else 0) + 1
    logger.debug("Assigning new defect ID '%s%d' (prefix: '%s')", prefix, max_int, prefix)
    defect_id = config.id_prefix + str(max_int).rjust(config.defect_id_digit_numbers, "0")

    defect_info_data_frame = create_defect_data_frame(defect, config, defect_id, protocol)

    return pd.concat([df, defect_info_data_frame], ignore_index=True)


def create_defect_data_frame(
    defect: Defect, config: ExcelDefectClientConfig, defect_id: str, protocol: Protocol
):

    defect_info_data_frame = pd.DataFrame(
        {
            "id": [defect_id],
            "title": [defect.title],
            "description": [defect.description],
            "reporter": [defect.reporter],
            "status": [defect.status],
            "classification": [defect.classification],
            "priority": [defect.priority],
            "lastEdited": [_format_last_edited(defect.lastEdited, config)],
            "references": [
                config.references_separator.join(defect.references if defect.references else [])
            ],
        }
    )

    if defect.userDefinedFields:
        for udf in defect.userDefinedFields:
            defect_info_data_frame[udf.name] = udf.value

    return defect_info_data_frame


def _format_last_edited(
    value: datetime | None,
    config: ExcelDefectClientConfig,
) -> str:
    if value is None:
        return ""
    format_string = to_python_datetime_format(config.simple_date_format)
    if format_string:
        return value.strftime(format_string)
    return value.isoformat()


def to_python_datetime_format(simple_date_format: str | None) -> str | None:
    if not simple_date_format:
        return None

    python_format = simple_date_format
    python_format = python_format.replace("yyyy", "%Y")
    python_format = python_format.replace("yy", "%y")
    python_format = python_format.replace("dd", "%d")
    python_format = python_format.replace("HH", "%H")
    python_format = python_format.replace("hh", "%I")
    python_format = python_format.replace("ss", "%S")
    python_format = python_format.replace("MM", "%m")
    if "HH" in simple_date_format or "hh" in simple_date_format:
        python_format = python_format.replace("mm", "%M")
    else:
        python_format = python_format.replace("mm", "%m")
    return python_format


def row_value(row: pd.Series, field_name: str) -> str:
    if field_name not in row.index:
        return ""
    value = row[field_name]
    return "" if value is None else str(value).strip()


def optional_row_value(row: pd.Series, field_name: str) -> str | None:
    value = row_value(row, field_name)
    return value or None


def is_blank_cell(value: Any) -> bool:
    """True when a cell holds no content: None, NaT, NaN, or whitespace only."""
    if value is None or value is pd.NaT:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    return str(value).strip() == ""


def is_blank_row(row: pd.Series) -> bool:
    """True when every mapped cell in the row is blank.

    An empty row in a defect file is layout, not data. The reader, the defect
    builder and the write mappers all have to agree on what counts as one.
    """
    return all(is_blank_cell(value) for value in row)


def split_references(raw_value: str, config: ExcelDefectClientConfig) -> list[str]:
    separator = config.references_separator or ";"
    if not raw_value:
        return []
    return [part.strip() for part in raw_value.split(separator) if part.strip()]


def parse_boolean_udf_value(
    raw_value: str,
    true_value: str | None,
    false_value: str | None,
) -> bool | None:
    normalized = raw_value.strip().lower()
    normalized_true = (true_value or "true").strip().lower()
    normalized_false = (false_value or "false").strip().lower()

    if normalized == normalized_true:
        return True
    if normalized == normalized_false:
        return False
    return None


def add_general_warning_once(
    protocol: Protocol,
    message: str,
    protocol_code: ProtocolCode,
) -> None:
    existing_warnings = protocol.generalWarnings or []
    if any(entry.message == message and entry.code == protocol_code for entry in existing_warnings):
        return
    protocol.add_general_warning(message, protocol_code=protocol_code)


def logical_field_names(sync_context: SyncContext) -> dict[str, str]:
    """Map each configured sync attribute name to the logical defect field it feeds."""
    mapping = {
        sync_context.statusAttribute: "status",
        sync_context.priorityAttribute: "priority",
        sync_context.classAttribute: "classification",
    }
    return {name: logical for name, logical in mapping.items() if name is not None}


def validate_control_fields(
    defect: Defect,
    config: ExcelDefectClientConfig,
    sync_context: SyncContext,
    protocol: Protocol,
) -> bool:
    logical_names = logical_field_names(sync_context)
    required_attributes = set(logical_names)
    validated: set[str] = set()

    for control_field in config.control_fields:
        logical = logical_names.get(control_field.name)
        if logical is None:
            continue
        value = getattr(defect, logical)
        if value not in control_field.values:
            protocol.add_error(
                key=control_field.name,
                message=(
                    f"Value '{value}' is not a valid option for control field "
                    f"'{control_field.name}'. Allowed values: {control_field.values}."
                ),
                protocol_code=ProtocolCode.PUBLISH_ERROR,
            )
            return False
        validated.add(control_field.name)

    missing_attributes = required_attributes - validated
    if missing_attributes:
        # Without this the caller aborts with an empty protocol, leaving TestBench to report
        # a failed defect with no reason attached.
        for attribute_name in sorted(missing_attributes):
            message = (
                f"Sync attribute '{attribute_name}' is not configured as a control field in "
                "the Excel client configuration."
            )
            logger.warning(message)
            protocol.add_error(
                key=attribute_name,
                message=message,
                protocol_code=ProtocolCode.PUBLISH_ERROR,
            )
        return False

    return True


def check_defect_transitions(
    defect: Defect,
    df: pd.DataFrame,
    config: ExcelDefectClientConfig,
    protocol: Protocol,
) -> bool:
    if defect.status == df["status"].values[0]:
        return True

    if config.transitions:
        for transition in config.transitions:
            if (
                transition.from_state == df["status"].values[0]
                and transition.to_state == defect.status
            ):
                return True
        current_status = df["status"].values[0]
        protocol.add_warning(
            key=defect.status,
            message=(
                f"No valid transition from '{current_status}' to '{defect.status}' is configured."
            ),
            protocol_code=ProtocolCode.INSERT_WARNING,
        )

        return False

    return True
